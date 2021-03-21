import urllib3
from urllib.parse import urlencode
from openpyxl import load_workbook
from common.ApiKey import ApiKey
from util.Common import convertUnixTime, AbTemperatureConvertCelsius
import json
import time
import datetime

def Load_Korea_latitude():
    #load_wb = load_workbook("/Korea_latitude&longtitude.xlsx")
    #load_ws = load_wb['Sheet1']
    #print(load_ws['D3'].value)
    #print(load_ws.cells(4,2).value)

    wb = load_workbook('util/Korea_latitude&longtitude.xlsx')
    ws = wb.active

    for r in ws.rows:
        row_index = r[1].row
        top = r[2].value
        mid = r[3].value
        bootom = r[4].value
        full = top+mid+bootom

        if u'서울' in full:
            print(full)
            break

        # print(type(row_index))
        # if len(str(row_index)) > 0:
        #     print(full)
        #     print("row = " + str(row_index))
        # if u'서울' in full:
        #     print("일치")
        #     print(str(row_index)+'@'+full)
    

    str1 = u'서울용두 동  날씨'
    str2 = u'서 울  용두동날씨'
    str3 = u'서울용두동날  씨'
    str4 = u'서울용두동  날씨'
    str5 = u'서울  용두동날씨'
    str6 = u'서울 용두동  날씨'

    strarray = [str1,str2,str3,str4,str5,str6]

    for str in strarray:
        print(str)
        str = str.replace(' ','')
        print(str)
        if '날씨' in str:
            print('있네있어')
        else:
            print("??")

def getLocalGeo(searchWord):
    http = urllib3.PoolManager()
    url = 'https://dapi.kakao.com/v2/local/search/address.json'
    query = '?' + urlencode({'query': searchWord})
    url+=query
    r = http.request('GET', url, headers={'Authorization':"KakaoAK %s" %ApiKey.KAKAO_KEY })
    data = json.loads(r.data.decode('UTF-8'))

    cnt = data['meta']['total_count']

    x = data['documents'][0]['x'] if cnt > 0 else 0 #latitude
    y = data['documents'][0]['y'] if cnt > 0 else 0 #longitude 

    return [y, x]

def getWeatherInfo(lat, lon):
    '''doc
    https://openweathermap.org/api/one-call-api
    '''

    http = urllib3.PoolManager()
    url = 'https://api.openweathermap.org/data/2.5/onecall'

    query = '?' + urlencode({'lat': lat, 'lon': lon, 'exclude':'minutely,hourly', 'lang':'kr', 'appid': ApiKey.OPENWEATHER_KEY})
    url+=query

    r = http.request('GET', url)
    data = json.loads(r.data.decode('UTF-8'))
    #print(data['current'])
    current = data['current']

    unix_time = int(current['dt'])
    #utc_time = time.gmtime(unix_time)
    local_time = time.localtime(unix_time)

    print(time.strftime('%Y-%m-%d %H:%M:%S', local_time))

    for daily in data['daily']:
        print(convertUnixTime(daily['dt']))
        print("아침:", round(float(daily['temp']['morn']) - 273.15))
        print("오후:", round(float(daily['temp']['day']) - 273.15))
        print("저녁:", round(float(daily['temp']['eve']) - 273.15))
        print("밤:", round(float(daily['temp']['night']) - 273.15))

        print("최저기온:", round(float(daily['temp']['min']) - 273.15))
        print("최고기온:", round(float(daily['temp']['max']) - 273.15))

        print("날씨요약:", daily['weather'][0]['description'])

        # current.temp
        # current.feels_like    
        # current.clouds 흐림
        # current.uvi UV지수
        # current.wind_speed 풍속
        # current.weather.description 날씨 상태 e.g. light snow
        # current.weather.icon e.g. 01d, http://openweathermap.org/img/wn/01d@2x.png

        # hourly 시간별 예보
        # hourly.dt 예측 시간 Unix Time
        # hourly.temp
        # hourly.feels_like
        # hourly.clouds
        # hourly.uvi
        # hourly.wind_speed
        # hourly.weather.description
        # hourly.weather.icon

        # daily.temp.morn 아침 온도
        # daily.temp.day 낮 온도
        # daily.temp.eve 저녁 온도
        # daily.temp.night 밤 온도
        # daily.temp.min 최소 일일 온도
        # daily.temp.max 최대 일일 온도
    return [data['current'], data['daily']]

def getWeatherInfo5days(lat, lon):
    '''doc
    https://openweathermap.org/forecast5
    '''
    
    http = urllib3.PoolManager()
    url = 'https://api.openweathermap.org/data/2.5/forecast'

    query = '?' + urlencode({'lat': lat, 'lon': lon, 'appid': ApiKey.OPENWEATHER_KEY})
    url+=query

    r = http.request('GET', url)
    data = json.loads(r.data.decode('UTF-8'))

    return data['list']


def App():
    # http = urllib3.PoolManager()
    # url = 'http://apis.data.go.kr/1360000/AsosHourlyInfoService/getWthrDataList'

    # serviceKey = ApiKey.OPENWEATHER_KEY

    # print(urlencode({'serviceKey': serviceKey}))

    # queryParams = '?' + urlencode({'serviceKey': serviceKey, 'pageNo': '1', 'numOfRows': '10', 'dataType': 'JSON', 'dataCd': 'ASOS', 'dateCd': 'HR', 
    #                                 'startDt': '20210130', 'startHh': '01', 'endDt': '20210130', 'endHh': '23', 'stnIds': '108' })
    # url+=queryParams

    # r = http.request('GET', url)

    # print(r.status)
    # print(r.data)

    #print(lon, lat)

    #Load_Korea_latitude()

    '''test
    lat, lon = getLocalGeo("자양동")
    getWeatherInfo(lat, lon)
    '''
#App()